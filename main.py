from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import JSONResponse, FileResponse, Response
from fastapi.middleware.cors import CORSMiddleware
import json
import os
import uuid
import qrcode
import io
from datetime import datetime
import csv
import base64
import google.generativeai as genai
from dotenv import load_dotenv
from face_utils import encode_face, compare_faces, save_encoding, load_encoding

# Load environment variables from .env file in the backend directory
env_path = os.path.join(os.path.dirname(__file__), '.env')
load_dotenv(env_path)

app = FastAPI()

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Data file paths
DATA_FILE = "data/data.json"
ENCODINGS_DIR = "data/encodings"
LOG_FILE = "data/recognition_log.csv"

# Gemini API configuration
# Load API key from .env file or environment variable
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
    print("Gemini API key loaded successfully")
else:
    print("Warning: GEMINI_API_KEY not found in .env file or environment variables")

def load_data():
    """Load user data from JSON file."""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_data(data):
    """Save user data to JSON file."""
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def log_recognition(user_id, name, class_name, roll, match_status, nutrition_data=None):
    """Log recognition event to CSV."""
    timestamp = datetime.now().isoformat()
    
    # Check if file exists and has header
    file_exists = os.path.exists(LOG_FILE) and os.path.getsize(LOG_FILE) > 0
    
    with open(LOG_FILE, 'a', newline='') as f:
        writer = csv.writer(f)
        
        # Write header if file is new
        if not file_exists:
            writer.writerow([
                'timestamp', 'user_id', 'name', 'class', 'roll', 'match_status',
                'carbohydrates', 'proteins', 'vitamins', 'fat', 'calories', 'other_nutrients'
            ])
        
        # Prepare nutrition values
        carbs = nutrition_data.get('carbohydrates', '') if nutrition_data else ''
        proteins = nutrition_data.get('proteins', '') if nutrition_data else ''
        vitamins = nutrition_data.get('vitamins', '') if nutrition_data else ''
        fat = nutrition_data.get('fat', '') if nutrition_data else ''
        calories = nutrition_data.get('calories', '') if nutrition_data else ''
        other = nutrition_data.get('other_nutrients', '') if nutrition_data else ''
        
        writer.writerow([
            timestamp, user_id, name, class_name, roll, match_status,
            carbs, proteins, vitamins, fat, calories, other
        ])

async def analyze_food_nutrition(image_bytes):
    """Analyze food image using Gemini API to detect nutrition."""
    if not GEMINI_API_KEY:
        print("Warning: GEMINI_API_KEY not set. Returning N/A values.")
        return {
            "error": "Gemini API key not configured. Please set GEMINI_API_KEY environment variable.",
            "carbohydrates": "N/A",
            "proteins": "N/A",
            "vitamins": "N/A",
            "fat": "N/A",
            "calories": "N/A",
            "other_nutrients": "N/A"
        }
    
    try:
        # Convert image to base64
        image_base64 = base64.b64encode(image_bytes).decode('utf-8')
        
        # Initialize Gemini model (using 1.5-flash as stable version)
        try:
            model = genai.GenerativeModel('gemini-2.5-flash')
        except:
            # Fallback to experimental version if available
            try:
                model = genai.GenerativeModel('gemini-2.0-flash-exp')
            except:
                model = genai.GenerativeModel('gemini-2.5-pro')
        
        # Create prompt for nutrition analysis
        prompt = """Analyze this food image and provide detailed nutritional information. 
        
        IMPORTANT: You must respond with ONLY a valid JSON object, no additional text or explanation.
        
        Identify and estimate the nutritional content including:
        - carbohydrates (in grams, e.g., "45g")
        - proteins (in grams, e.g., "20g")
        - vitamins (list the vitamins present, e.g., "Vitamin A, Vitamin C, Vitamin B12")
        - fat (in grams, e.g., "15g")
        - calories (estimated number, e.g., "350")
        - other_nutrients (any other important nutrients, e.g., "Fiber: 5g, Iron: 2mg")
        
        Return ONLY a JSON object with these exact keys: carbohydrates, proteins, vitamins, fat, calories, other_nutrients.
        If you cannot determine a value, use "N/A" for that field.
        
        Example response (return ONLY this format, no markdown, no code blocks):
        {"carbohydrates": "45g", "proteins": "20g", "vitamins": "Vitamin A, Vitamin C", "fat": "15g", "calories": "350", "other_nutrients": "Fiber: 5g"}"""
        
        # Prepare the image part
        from PIL import Image
        image = Image.open(io.BytesIO(image_bytes))
        
        # Generate content
        response = model.generate_content([prompt, image])
        
        # Parse response
        response_text = response.text.strip()
        print(f"Gemini response: {response_text}")  # Debug log
        
        # Try to extract JSON from response (handle markdown code blocks)
        import re
        
        # Remove markdown code blocks if present
        response_text = re.sub(r'```json\s*', '', response_text)
        response_text = re.sub(r'```\s*', '', response_text)
        response_text = response_text.strip()
        
        # Try multiple strategies to extract JSON
        nutrition_data = None
        
        # Strategy 1: Try parsing the whole response as JSON
        try:
            nutrition_data = json.loads(response_text)
        except json.JSONDecodeError:
            pass
        
        # Strategy 2: Find JSON object using balanced braces
        if nutrition_data is None:
            brace_count = 0
            start_idx = -1
            for i, char in enumerate(response_text):
                if char == '{':
                    if brace_count == 0:
                        start_idx = i
                    brace_count += 1
                elif char == '}':
                    brace_count -= 1
                    if brace_count == 0 and start_idx != -1:
                        json_str = response_text[start_idx:i+1]
                        try:
                            nutrition_data = json.loads(json_str)
                            break
                        except json.JSONDecodeError:
                            continue
        
        # Strategy 3: Try regex for simple JSON
        if nutrition_data is None:
            json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text, re.DOTALL)
            if json_match:
                try:
                    nutrition_data = json.loads(json_match.group())
                except json.JSONDecodeError:
                    pass
        
        # Fallback: create default structure
        if nutrition_data is None:
            print(f"Warning: Could not parse JSON from response: {response_text[:200]}")
            nutrition_data = {
                "carbohydrates": "N/A",
                "proteins": "N/A",
                "vitamins": "N/A",
                "fat": "N/A",
                "calories": "N/A",
                "other_nutrients": "N/A"
            }
        
        # Ensure all required keys exist
        required_keys = ["carbohydrates", "proteins", "vitamins", "fat", "calories", "other_nutrients"]
        for key in required_keys:
            if key not in nutrition_data:
                nutrition_data[key] = "N/A"
        
        return nutrition_data
    
    except Exception as e:
        print(f"Error analyzing food nutrition: {e}")
        return {
            "error": f"Failed to analyze food image: {str(e)}",
            "carbohydrates": "N/A",
            "proteins": "N/A",
            "vitamins": "N/A",
            "fat": "N/A",
            "calories": "N/A",
            "other_nutrients": "N/A"
        }

def generate_qr_code(user_id):
    """Generate QR code image for user_id."""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(user_id)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to bytes
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    
    return img_bytes

@app.post("/api/register")
async def register_user(
    name: str = Form(...),
    class_name: str = Form(...),
    roll: str = Form(...),
    aadhar: str = Form(None),
    face_image: UploadFile = File(...)
):
    """Register a new user with face encoding."""
    try:
        # Read image bytes
        image_bytes = await face_image.read()
        
        # Encode face
        encoding = encode_face(image_bytes)
        
        if encoding is None:
            raise HTTPException(status_code=400, detail="No face detected in the image. Please ensure your face is clearly visible.")
        
        # Generate unique user ID
        user_id = str(uuid.uuid4())
        
        # Save face encoding
        if not save_encoding(user_id, encoding, ENCODINGS_DIR):
            raise HTTPException(status_code=500, detail="Failed to save face encoding")
        
        # Load existing data
        data = load_data()
        
        # Save user data
        data[user_id] = {
            "name": name,
            "class": class_name,
            "roll": roll,
            "aadhar": aadhar if aadhar else None,
            "registration_date": datetime.now().isoformat()
        }
        save_data(data)
        
        # Generate QR code
        qr_image = generate_qr_code(user_id)
        
        # Save QR code temporarily (or return as base64)
        qr_path = os.path.join(ENCODINGS_DIR, f"{user_id}_qr.png")
        with open(qr_path, 'wb') as f:
            f.write(qr_image.getvalue())
        
        return JSONResponse({
            "success": True,
            "user_id": user_id,
            "message": "User registered successfully",
            "qr_path": f"/api/qr/{user_id}"
        })
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Registration failed: {str(e)}")

@app.post("/api/recognize")
async def recognize_user(
    user_id: str = Form(...),
    face_image: UploadFile = File(...)
):
    """Recognize a user by comparing face with stored encoding."""
    try:
        # Load user data
        data = load_data()
        
        if user_id not in data:
            raise HTTPException(status_code=404, detail="User not found")
        
        user_data = data[user_id]
        
        # Load stored encoding
        stored_encoding = load_encoding(user_id, ENCODINGS_DIR)
        
        if stored_encoding is None:
            raise HTTPException(status_code=404, detail="Face encoding not found for this user")
        
        # Read and encode new image
        image_bytes = await face_image.read()
        new_encoding = encode_face(image_bytes)
        
        if new_encoding is None:
            raise HTTPException(status_code=400, detail="No face detected in the image. Please ensure your face is clearly visible.")
        
        # Compare faces
        match = compare_faces(stored_encoding, new_encoding)
        
        # Log recognition event
        log_recognition(
            user_id,
            user_data["name"],
            user_data["class"],
            user_data["roll"],
            "matched" if match else "not_matched"
        )
        
        if match:
            return JSONResponse({
                "success": True,
                "match": True,
                "message": "Face recognized successfully",
                "user_data": user_data
            })
        else:
            return JSONResponse({
                "success": False,
                "match": False,
                "message": "Face does not match. Please try again."
            })
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Recognition failed: {str(e)}")

@app.post("/api/analyze-food")
async def analyze_food(
    user_id: str = Form(...),
    food_image: UploadFile = File(...)
):
    """Analyze food image and update recognition log with nutrition data."""
    try:
        # Load user data
        data = load_data()
        
        if user_id not in data:
            raise HTTPException(status_code=404, detail="User not found")
        
        user_data = data[user_id]
        
        # Read food image
        food_image_bytes = await food_image.read()
        
        # Analyze nutrition using Gemini
        nutrition_data = await analyze_food_nutrition(food_image_bytes)
        
        # Handle errors - ensure we always have valid nutrition data structure
        if "error" in nutrition_data:
            if not GEMINI_API_KEY:
                # If API key not set, use N/A values
                nutrition_data = {
                    "carbohydrates": "N/A",
                    "proteins": "N/A",
                    "vitamins": "N/A",
                    "fat": "N/A",
                    "calories": "N/A",
                    "other_nutrients": "N/A"
                }
            else:
                # API key is set but there was an error - still use the structure if available
                if not all(key in nutrition_data for key in ["carbohydrates", "proteins", "vitamins", "fat", "calories", "other_nutrients"]):
                    nutrition_data = {
                        "carbohydrates": nutrition_data.get("carbohydrates", "N/A"),
                        "proteins": nutrition_data.get("proteins", "N/A"),
                        "vitamins": nutrition_data.get("vitamins", "N/A"),
                        "fat": nutrition_data.get("fat", "N/A"),
                        "calories": nutrition_data.get("calories", "N/A"),
                        "other_nutrients": nutrition_data.get("other_nutrients", "N/A")
                    }
        
        # Update the last recognition log entry for this user
        # Read existing CSV
        rows = []
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)
        
        # Update header if needed
        if len(rows) > 0:
            header = rows[0]
            expected_header = ['timestamp', 'user_id', 'name', 'class', 'roll', 'match_status',
                              'carbohydrates', 'proteins', 'vitamins', 'fat', 'calories', 'other_nutrients']
            if len(header) < len(expected_header):
                rows[0] = expected_header
        
        # Find and update the last entry for this user_id
        updated = False
        if len(rows) > 1:  # Has header + at least one data row
            # Search from bottom to top for the last entry with this user_id
            for i in range(len(rows) - 1, 0, -1):
                if len(rows[i]) > 1 and rows[i][1] == user_id:  # user_id is in column 1
                    # Update this row with nutrition data
                    # Ensure row has enough columns
                    while len(rows[i]) < 12:
                        rows[i].append('')
                    
                    rows[i][6] = nutrition_data.get('carbohydrates', 'N/A')
                    rows[i][7] = nutrition_data.get('proteins', 'N/A')
                    rows[i][8] = nutrition_data.get('vitamins', 'N/A')
                    rows[i][9] = nutrition_data.get('fat', 'N/A')
                    rows[i][10] = nutrition_data.get('calories', 'N/A')
                    rows[i][11] = nutrition_data.get('other_nutrients', 'N/A')
                    updated = True
                    break
        
        # If no matching entry found, create a new one
        if not updated:
            timestamp = datetime.now().isoformat()
            new_row = [
                timestamp, user_id, user_data["name"], user_data["class"], 
                user_data["roll"], "food_analyzed",
                nutrition_data.get('carbohydrates', 'N/A'),
                nutrition_data.get('proteins', 'N/A'),
                nutrition_data.get('vitamins', 'N/A'),
                nutrition_data.get('fat', 'N/A'),
                nutrition_data.get('calories', 'N/A'),
                nutrition_data.get('other_nutrients', 'N/A')
            ]
            rows.append(new_row)
        
        # Write back to CSV
        with open(LOG_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        
        # Check if there was an error in analysis
        has_error = "error" in nutrition_data and GEMINI_API_KEY
        error_message = nutrition_data.get("error", "") if has_error else None
        
        return JSONResponse({
            "success": True,
            "message": "Food analyzed successfully" if not has_error else f"Food analyzed with warnings: {error_message}",
            "nutrition_data": nutrition_data,
            "warning": error_message if has_error else None
        })
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Food analysis failed: {str(e)}")

@app.get("/api/qr/{user_id}")
async def get_qr_code(user_id: str):
    """Get QR code image for a user."""
    try:
        # Check if user exists
        data = load_data()
        if user_id not in data:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Check if QR code file exists
        qr_path = os.path.join(ENCODINGS_DIR, f"{user_id}_qr.png")
        
        if not os.path.exists(qr_path):
            # Generate QR code if it doesn't exist
            qr_image = generate_qr_code(user_id)
            with open(qr_path, 'wb') as f:
                f.write(qr_image.getvalue())
        
        return FileResponse(qr_path, media_type="image/png", filename=f"{user_id}_qr.png")
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate QR code: {str(e)}")

@app.get("/api/user/{user_id}")
async def get_user(user_id: str):
    """Get user data by user_id."""
    try:
        data = load_data()
        
        if user_id not in data:
            raise HTTPException(status_code=404, detail="User not found")
        
        return JSONResponse({
            "success": True,
            "user_data": data[user_id]
        })
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to retrieve user: {str(e)}")

@app.get("/api/logs")
async def get_logs():
    """Get all recognition logs from CSV."""
    try:
        logs = []
        
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', newline='') as f:
                reader = csv.DictReader(f)
                logs = list(reader)
        
        return JSONResponse({
            "success": True,
            "logs": logs,
            "count": len(logs)
        })
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to retrieve logs: {str(e)}")

@app.get("/api/logs/excel")
async def download_logs_excel():
    """Download recognition logs as Excel file."""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="openpyxl is not installed. Please run: pip install openpyxl"
            )
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Recognition Logs"
        
        # Read CSV data
        logs = []
        headers = []
        
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', newline='') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                logs = [row for row in reader if row.get('timestamp', '').strip()]
        
        if not headers:
            headers = ['timestamp', 'user_id', 'name', 'class', 'roll', 'match_status',
                       'carbohydrates', 'proteins', 'vitamins', 'fat', 'calories', 'other_nutrients']
        
        # Write headers
        header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header.replace('_', ' ').title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_num, log in enumerate(logs, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = log.get(header, 'N/A')
                
                # Format timestamp if it's a timestamp column
                if header == 'timestamp' and value and value != 'N/A':
                    try:
                        from datetime import datetime
                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        cell.value = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        cell.value = value
                else:
                    cell.value = value
                
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        if ws.max_row > 1:  # Only adjust if there's data
            for col_num, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                    if row[0].value:
                        max_length = max(max_length, len(str(row[0].value)))
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        else:
            # Set default widths if no data
            for col_num, header in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = len(str(header)) + 2
        
        # Set header row height
        ws.row_dimensions[1].height = 25
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"recognition_logs_{timestamp}.xlsx"
        
        # Use Response for in-memory file instead of FileResponse
        return Response(
            content=excel_bytes.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel generation error: {error_details}")
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel file: {str(e)}")

@app.get("/")
async def root():
    """Root endpoint."""
    return {"message": "Face Recognition API"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

